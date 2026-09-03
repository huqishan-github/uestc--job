#!/usr/bin/env python3
"""Incrementally monitor public recruitment information on the UESTC career site."""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import logging
import os
import re
import tempfile
import threading
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence
from urllib.parse import parse_qsl, urlencode, urljoin, urlsplit, urlunsplit
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CONFIG_PATH = ROOT / "config" / "config.json"
CSV_PATH = ROOT / "data" / "jobs.csv"
EXCEL_PATH = ROOT / "data" / "UESTC招聘信息.xlsx"
REPORTS_DIR = ROOT / "reports"

CSV_FIELDS = [
    "唯一ID",
    "首次获取时间",
    "最后更新时间",
    "信息类型",
    "公司名称",
    "招聘标题",
    "岗位",
    "发布时间",
    "招聘日期",
    "工作地点",
    "宣讲日期",
    "宣讲开始时间",
    "宣讲结束时间",
    "宣讲地点",
    "学历要求",
    "专业要求",
    "网申截止时间",
    "投递方式",
    "投递链接",
    "详情链接",
    "链接状态",
    "来源页面",
    "备注",
]

EXCEL_FIELDS = [
    "唯一ID",
    "首次获取时间",
    "信息类型",
    "公司名称",
    "招聘标题",
    "岗位",
    "发布时间",
    "招聘日期",
    "工作地点",
    "宣讲日期",
    "宣讲开始时间",
    "宣讲结束时间",
    "宣讲地点",
    "学历要求",
    "专业要求",
    "网申截止时间",
    "投递方式",
    "投递链接",
    "详情链接",
    "链接状态",
    "来源页面",
    "最后更新时间",
    "备注",
]

UPDATE_FIELDS = ["时间", "操作类型", "唯一ID", "公司名称", "岗位", "变化说明"]
IMMUTABLE_FIELDS = {"唯一ID", "首次获取时间", "最后更新时间"}
URL_FIELDS = {"投递链接", "详情链接", "来源页面"}
TRACKING_QUERY_KEYS = {
    "gclid",
    "fbclid",
    "yclid",
    "mc_cid",
    "mc_eid",
    "spm",
    "from",
}
LINK_STATUS_RANK = {"": 0, "failed": 1, "unverified": 2, "verified": 3}
USER_AGENT = "UESTC-Job-Monitor/1.0 (+low-frequency public recruitment monitoring)"
ONSITE_ROUTE_SLUGS = {
    "ONSITE": "onsite",
    "GROUP": "group",
    "RECRUITMENT": "aerial",
}

JOB_HEADER_NAMES = {
    "岗位",
    "职位",
    "岗位名称",
    "职位名称",
    "招聘岗位",
    "招聘职位",
    "需求岗位",
    "应聘岗位",
    "岗位类别",
    "职位类别",
}
JOB_HEADER_PRIORITY = (
    "岗位名称",
    "职位名称",
    "招聘岗位",
    "招聘职位",
    "需求岗位",
    "应聘岗位",
    "岗位",
    "职位",
    "职位类别",
    "岗位类别",
)
MAJOR_HEADER_NAMES = {"专业", "专业要求", "所需专业", "招聘专业", "需求专业", "专业方向"}
EDUCATION_HEADER_NAMES = {"学历", "学历要求", "最低学历", "学位要求"}
LOCATION_HEADER_NAMES = {"工作地点", "工作地", "工作地址", "岗位地点", "办公地点"}
DEADLINE_HEADER_NAMES = {"截止时间", "截止日期", "网申截止时间", "网申截止", "申请截止时间"}
JOB_SUFFIX_RE = re.compile(
    r"(?:工程师|设计师|研究员|分析师|专员|经理|顾问|管培生|技术员|助理|教师|博士后|"
    r"客户经理|产品经理|项目经理|销售|运营|开发|测试|算法|研发|岗)$"
)
ROLE_ENDING_RE = re.compile(
    r"(?:工程师|设计师|研究员|分析师|专员|经理|顾问|管培生|技术员|助理|教师|博士后|"
    r"设计|开发|控制|处理|仿真|测试|规划|管理|支持|采购|算法|工艺|销售)$"
)
STRONG_ROLE_BOUNDARY_RE = re.compile(
    r"(?:工程师|设计师|研究员|分析师|专员|经理|顾问|管培生|技术员|助理|教师|博士后|"
    r"销售(?=(?:技术|数据|售后|硬件|软件|算法|产品|项目|研发|采购)))"
    r"(?=[A-Za-z0-9\u4e00-\u9fff])"
)


class MonitorError(RuntimeError):
    """Base class for controlled monitor failures."""


class FetchError(MonitorError):
    """Raised when a network request cannot be completed safely."""


class DataFileError(MonitorError):
    """Raised when existing history cannot be read without risking data loss."""


@dataclass(frozen=True)
class Candidate:
    """One announcement discovered in a public list/calendar response."""

    source_type: str
    source_code: str
    source_slug: str
    source_page: str
    raw: Mapping[str, Any]

    @property
    def announcement_id(self) -> str:
        return normalize_text(self.raw.get("id"))


@dataclass
class Position:
    name: str
    education: str = ""
    major: str = ""
    location: str = ""
    deadline: str = ""


@dataclass
class DetailResult:
    candidate: Candidate
    detail_url: str
    link_status: str
    records: list[dict[str, str]] = field(default_factory=list)
    error: str = ""


@dataclass
class ChangeEvent:
    timestamp: str
    operation: str
    unique_id: str
    company: str
    job: str
    description: str


@dataclass
class MergeResult:
    rows: list[dict[str, str]]
    new_records: list[dict[str, str]]
    updated_records: list[dict[str, str]]
    events: list[ChangeEvent]
    unchanged: int


@dataclass
class SourceRun:
    label: str
    discovered: int = 0
    selected: int = 0
    parsed_rows: int = 0
    new: int = 0
    updated: int = 0
    unchanged: int = 0
    verified: int = 0
    unverified: int = 0
    failed: int = 0
    discovery_error: str = ""


def normalize_text(value: Any) -> str:
    """Normalize Unicode and whitespace without inventing or translating content."""

    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        value = "、".join(normalize_text(item) for item in value if normalize_text(item))
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\u200b", "").replace("\ufeff", "")
    return re.sub(r"\s+", " ", text).strip()


def normalize_url(value: Any, base_url: str | None = None) -> str:
    """Canonicalize a real URL and remove only obvious tracking parameters."""

    raw = normalize_text(value)
    if not raw:
        return ""
    if base_url:
        raw = urljoin(base_url, raw)
    try:
        parts = urlsplit(raw)
    except ValueError:
        return raw.rstrip("/")
    if parts.scheme.lower() not in {"http", "https"} or not parts.netloc:
        return raw.rstrip("/")
    scheme = parts.scheme.lower()
    netloc = parts.netloc.lower()
    if scheme == "http" and netloc.endswith(":80"):
        netloc = netloc[:-3]
    elif scheme == "https" and netloc.endswith(":443"):
        netloc = netloc[:-4]
    path = parts.path or "/"
    if path != "/":
        path = path.rstrip("/")
    kept_query = []
    for key, val in parse_qsl(parts.query, keep_blank_values=True):
        lowered = key.lower()
        if lowered.startswith("utm_") or lowered in TRACKING_QUERY_KEYS:
            continue
        kept_query.append((key, val))
    kept_query.sort()
    return urlunsplit((scheme, netloc, path, urlencode(kept_query, doseq=True), ""))


def make_unique_id(record: Mapping[str, Any]) -> str:
    """Build the stable SHA-256 ID defined by the project data contract."""

    source_type = normalize_text(record.get("信息类型"))
    detail_url = normalize_url(record.get("详情链接"))
    job_name = normalize_text(record.get("岗位"))
    has_explicit_job = bool(job_name and job_name != "未明确列出")
    if detail_url:
        parts = [source_type, detail_url]
        if has_explicit_job:
            parts.append(job_name)
    else:
        parts = [
            source_type,
            normalize_text(record.get("公司名称")),
            normalize_text(record.get("招聘标题")),
            normalize_text(record.get("发布时间") or record.get("招聘日期")),
            job_name,
        ]
    material = "\x1f".join(parts)
    return hashlib.sha256(material.encode("utf-8")).hexdigest()


def encode_public_id(value: str) -> str:
    """Apply the base-36 route encoding used by the site's own Nuxt client."""

    value = normalize_text(value)
    if not value.isdigit():
        return f"-{value}-"
    number = int(value)
    alphabet = "0123456789abcdefghijklmnopqrstuvwxyz"
    if number == 0:
        return "0"
    chars: list[str] = []
    while number:
        number, remainder = divmod(number, 36)
        chars.append(alphabet[remainder])
    return "".join(reversed(chars))


def parse_datetime(value: Any) -> datetime | None:
    text = normalize_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            continue
    return None


def format_datetime(value: Any) -> str:
    parsed = parse_datetime(value)
    return parsed.strftime("%Y-%m-%d %H:%M:%S") if parsed else normalize_text(value)


def format_date(value: Any) -> str:
    parsed = parse_datetime(value)
    return parsed.strftime("%Y-%m-%d") if parsed else normalize_text(value)


def format_time(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    match = re.search(r"(?:^|\s)(\d{1,2}:\d{2})(?::\d{2})?", text)
    return match.group(1) if match else text


def join_values(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.startswith("["):
            try:
                value = json.loads(stripped)
            except (json.JSONDecodeError, TypeError):
                return normalize_text(value)
        else:
            return normalize_text(value)
    if isinstance(value, (list, tuple, set)):
        result: list[str] = []
        for item in value:
            normalized = normalize_text(item)
            if normalized and normalized not in result:
                result.append(normalized)
        return "、".join(result)
    return normalize_text(value)


class RateLimiter:
    """Limit request start times across all detail-fetch worker threads."""

    def __init__(self, interval_seconds: float) -> None:
        self.interval = max(0.0, interval_seconds)
        self._next_start = 0.0
        self._lock = threading.Lock()

    def wait(self) -> None:
        with self._lock:
            now = time.monotonic()
            start_at = max(now, self._next_start)
            self._next_start = start_at + self.interval
        delay = start_at - now
        if delay > 0:
            time.sleep(delay)


class UestcClient:
    """Small, retry-bounded client for the public UESTC career APIs and pages."""

    def __init__(self, config: Mapping[str, Any]) -> None:
        self.base_url = normalize_url(config["base_url"]) + "/"
        self.timeout = float(config["request_timeout"])
        self.max_retries = int(config["max_retries"])
        self.limiter = RateLimiter(float(config["request_interval_seconds"]))
        self._local = threading.local()

    def _session(self) -> requests.Session:
        session = getattr(self._local, "session", None)
        if session is None:
            session = requests.Session()
            session.headers.update({"User-Agent": USER_AGENT, "Accept": "*/*"})
            self._local.session = session
        return session

    def request(self, method: str, url: str, **kwargs: Any) -> requests.Response:
        last_error = ""
        for attempt in range(1, self.max_retries + 1):
            self.limiter.wait()
            try:
                response = self._session().request(
                    method,
                    url,
                    timeout=self.timeout,
                    allow_redirects=True,
                    **kwargs,
                )
            except requests.RequestException as exc:
                last_error = f"{exc.__class__.__name__}: {exc}"
            else:
                if response.status_code < 400:
                    return response
                last_error = f"HTTP {response.status_code}"
                retryable = response.status_code in {403, 408, 429} or response.status_code >= 500
                if not retryable:
                    raise FetchError(f"{url} 返回 {last_error}")
            if attempt < self.max_retries:
                time.sleep(min(2 ** (attempt - 1), 4))
        raise FetchError(f"{url} 请求失败（重试 {self.max_retries} 次）：{last_error}")

    def api_get(self, path: str, params: Mapping[str, Any] | None = None) -> Mapping[str, Any]:
        url = urljoin(self.base_url, path.lstrip("/"))
        response = self.request("GET", url, params=params)
        return self._parse_api_response(response, url)

    def api_post(self, path: str, payload: Mapping[str, Any]) -> Mapping[str, Any]:
        url = urljoin(self.base_url, path.lstrip("/"))
        response = self.request("POST", url, json=payload, headers={"Content-Type": "application/json"})
        return self._parse_api_response(response, url)

    @staticmethod
    def _parse_api_response(response: requests.Response, url: str) -> Mapping[str, Any]:
        try:
            payload = response.json()
        except ValueError as exc:
            raise FetchError(f"{url} 未返回有效 JSON") from exc
        if not isinstance(payload, dict) or not payload.get("ok"):
            code = payload.get("code") if isinstance(payload, dict) else "unknown"
            raise FetchError(f"{url} API 返回异常，code={code}")
        return payload


def months_between(start: date, end: date) -> list[str]:
    result: list[str] = []
    cursor = date(start.year, start.month, 1)
    last = date(end.year, end.month, 1)
    while cursor <= last:
        result.append(cursor.strftime("%Y-%m"))
        cursor = date(cursor.year + (cursor.month == 12), 1 if cursor.month == 12 else cursor.month + 1, 1)
    return result


def discover_onsite(
    client: UestcClient,
    base_url: str,
    today: date,
    recent_days: int,
    forward_days: int,
) -> list[Candidate]:
    start = today - timedelta(days=recent_days)
    end = today + timedelta(days=forward_days)
    found: dict[str, Candidate] = {}
    for month in months_between(start, end):
        payload = client.api_get("api/home/calendar", {"month": month})
        calendar = payload.get("data", {}).get("onSiteCalendar", {})
        days = calendar.get("data", {}) if isinstance(calendar, dict) else {}
        if not isinstance(days, dict):
            raise FetchError(f"现场招聘日历 {month} 的数据结构已改变")
        for day_info in days.values():
            items = day_info.get("data", []) if isinstance(day_info, dict) else []
            for item in items if isinstance(items, list) else []:
                if not isinstance(item, dict):
                    continue
                event_dt = parse_datetime(item.get("recruitmentDate"))
                item_id = normalize_text(item.get("id"))
                source_code = normalize_text(item.get("recruitmentTypeCode"))
                source_slug = ONSITE_ROUTE_SLUGS.get(source_code)
                if event_dt and start <= event_dt.date() <= end and item_id and source_slug:
                    found[item_id] = Candidate(
                        source_type="现场招聘",
                        source_code=source_code,
                        source_slug=source_slug,
                        source_page=urljoin(base_url, f"recruitment/{source_slug}"),
                        raw=item,
                    )

    def order(candidate: Candidate) -> tuple[int, date]:
        event_dt = parse_datetime(candidate.raw.get("recruitmentDate"))
        event_date = event_dt.date() if event_dt else date.max
        if event_date >= today:
            return 0, event_date
        return 1, date.max - (event_date - date.min)

    return sorted(found.values(), key=order)


def discover_online(
    client: UestcClient,
    base_url: str,
    cutoff: datetime,
    page_size: int,
    max_pages: int,
) -> list[Candidate]:
    found: dict[str, Candidate] = {}
    for page_index in range(1, max_pages + 1):
        payload = client.api_post(
            "api/home/recruitmentList",
            {
                "type": "ONLINE_RECRUITMENT",
                "timeRange": "ALL",
                "pageIndex": page_index,
                "pageSize": page_size,
                "isSearchPage": False,
                "searchParams": [],
            },
        )
        items = payload.get("data", [])
        if not isinstance(items, list):
            raise FetchError("需求信息列表 API 的数据结构已改变")
        parsed_times: list[datetime] = []
        for item in items:
            if not isinstance(item, dict):
                continue
            published = parse_datetime(item.get("publishTime"))
            if published:
                parsed_times.append(published)
            item_id = normalize_text(item.get("id"))
            if item_id and (published is None and page_index == 1 or published and published >= cutoff):
                found[item_id] = Candidate(
                    source_type="需求信息",
                    source_code="ONLINE_RECRUITMENT",
                    source_slug="online",
                    source_page=urljoin(base_url, "recruitment/online"),
                    raw=item,
                )
        page = payload.get("page", {})
        total_pages = int(page.get("totalPage") or max_pages) if isinstance(page, dict) else max_pages
        is_descending = all(a >= b for a, b in zip(parsed_times, parsed_times[1:]))
        if not items or page_index >= total_pages:
            break
        if parsed_times and min(parsed_times) < cutoff and is_descending:
            break
    return sorted(
        found.values(),
        key=lambda candidate: parse_datetime(candidate.raw.get("publishTime")) or datetime.min,
        reverse=True,
    )


def _resolve_devalue(
    pool: Sequence[Any],
    value: Any,
    trail: frozenset[int] = frozenset(),
    reference: bool = True,
) -> Any:
    """Resolve the reference-array format used in Nuxt's __NUXT_DATA__ payload."""

    if reference and isinstance(value, int) and not isinstance(value, bool) and 0 <= value < len(pool):
        if value in trail:
            return None
        resolved = pool[value]
        if isinstance(resolved, (list, dict)):
            return _resolve_devalue(pool, resolved, trail | {value}, reference=False)
        return resolved
    if isinstance(value, list):
        if len(value) == 2 and value[0] in {"Ref", "Reactive", "ShallowReactive"}:
            return _resolve_devalue(pool, value[1], trail, reference=True)
        return [_resolve_devalue(pool, item, trail, reference=True) for item in value]
    if isinstance(value, dict):
        return {key: _resolve_devalue(pool, item, trail, reference=True) for key, item in value.items()}
    return value


def extract_nuxt_recruitment(page_html: str, expected_id: str) -> Mapping[str, Any] | None:
    soup = BeautifulSoup(page_html, "html.parser")
    payload_tag = soup.find("script", id="__NUXT_DATA__")
    if not payload_tag or not payload_tag.string:
        return None
    try:
        pool = json.loads(payload_tag.string)
    except (json.JSONDecodeError, TypeError):
        return None
    if not isinstance(pool, list):
        return None
    for item in pool:
        if not isinstance(item, dict):
            continue
        if not {"id", "companyName", "recruitmentTypeCode"}.issubset(item):
            continue
        resolved_id = _resolve_devalue(pool, item.get("id"))
        if normalize_text(resolved_id) == normalize_text(expected_id):
            return {key: _resolve_devalue(pool, value) for key, value in item.items()}
    return None


def extract_recruitment_title(content_html: str, api_title: Any = "") -> str:
    """Extract only an explicitly displayed announcement title."""

    soup = BeautifulSoup(content_html or "", "html.parser")
    candidates: list[str] = []
    for tag in soup.find_all(["h1", "h2", "h3", "h4"]):
        text = normalize_text(tag.get_text(" ", strip=True))
        if 4 <= len(text) <= 150 and text not in candidates:
            candidates.append(text)
    if not candidates:
        for tag in soup.find_all(["p", "div"], limit=60):
            style = normalize_text(tag.get("style")).lower()
            text = normalize_text(tag.get_text(" ", strip=True))
            if "center" in style and 4 <= len(text) <= 150 and text not in candidates:
                candidates.append(text)
    for candidate in candidates:
        if re.search(r"招聘|招募|校园|简章|公告", candidate):
            return candidate
    if candidates:
        return candidates[0]
    # The API title is displayed as a position by the site's own list component; use it
    # as an announcement title only when it explicitly looks like an announcement.
    api_text = normalize_text(api_title)
    return api_text if re.search(r"招聘|招募|简章|公告", api_text) else ""


def _header_key(text: str) -> str:
    return re.sub(r"[\s:：()（）/]+", "", normalize_text(text))


def _find_header(headers: Sequence[str], options: set[str]) -> int | None:
    normalized_options = {_header_key(option) for option in options}
    for index, header in enumerate(headers):
        key = _header_key(header)
        if key in normalized_options:
            return index
    return None


def _find_job_header(headers: Sequence[str]) -> int | None:
    normalized = [_header_key(header) for header in headers]
    for option in JOB_HEADER_PRIORITY:
        key = _header_key(option)
        if key in normalized:
            return normalized.index(key)
    return None


def _clean_job_name(value: str) -> str:
    value = normalize_text(value)
    value = re.sub(r"^(?:岗位|职位|招聘岗位|招聘职位|岗位名称|职位名称)\s*[：:]\s*", "", value)
    value = re.sub(r"^(?:\d+|[一二三四五六七八九十]+)[.、)）]\s*", "", value)
    return value.strip(" ：:;；、,，")


def _looks_like_job_name(value: str) -> bool:
    value = _clean_job_name(value)
    if not 2 <= len(value) <= 80:
        return False
    if value in JOB_HEADER_NAMES or value in {"合计", "总计", "详见公告", "详见招聘简章", "若干"}:
        return False
    if re.search(r"岗位职责|任职要求|专业要求|学历要求|工作地点|招聘人数|薪资|福利|联系方式", value):
        return False
    if re.search(r"[。！？?]", value):
        return False
    return True


def split_outside_brackets(value: str, separators: set[str], whitespace: bool = False) -> list[str]:
    """Split on visible separators while preserving text inside brackets."""

    opening = {"(", "[", "{", "（", "【", "［", "｛"}
    closing = {")", "]", "}", "）", "】", "］", "｝"}
    depth = 0
    parts: list[str] = []
    current: list[str] = []
    for char in value:
        if char in opening:
            depth += 1
        elif char in closing and depth:
            depth -= 1
        is_separator = depth == 0 and (char in separators or whitespace and char.isspace())
        if is_separator:
            part = "".join(current).strip()
            if part:
                parts.append(part)
            current = []
        else:
            current.append(char)
    tail = "".join(current).strip()
    if tail:
        parts.append(tail)
    return parts


def _role_stem(value: str) -> str:
    return re.sub(r"\s*[（(【\[].*$", "", normalize_text(value).removesuffix("等")).strip()


def _ends_like_role(value: str) -> bool:
    return bool(ROLE_ENDING_RE.search(_role_stem(value)))


def split_repeated_strong_roles(value: str) -> list[str]:
    """Split concatenated complete role names such as '算法工程师项目工程师'."""

    depth = 0
    boundary_ends: list[int] = []
    for index, char in enumerate(value):
        if char in "([{（【［｛":
            depth += 1
        elif char in ")]}）】］｝" and depth:
            depth -= 1
        if depth:
            continue
        match = STRONG_ROLE_BOUNDARY_RE.match(value, index)
        if match:
            boundary_ends.append(match.end())
    if not boundary_ends:
        return [value]
    pieces: list[str] = []
    start = 0
    for end in boundary_ends:
        piece = _clean_job_name(value[start:end])
        if piece:
            pieces.append(piece)
        start = end
    tail = _clean_job_name(value[start:])
    if tail:
        pieces.append(tail)
    if len(pieces) > 1 and all(
        _looks_like_job_name(piece) and _ends_like_role(piece) for piece in pieces
    ):
        return pieces
    return [value]


def split_job_names(value: str) -> list[str]:
    """Split only explicit lists; keep ambiguous compound role phrases intact."""

    text = unicodedata.normalize("NFKC", str(value or ""))
    text = text.replace("\r", "\n").replace("\u200b", "").replace("\ufeff", "")
    text = re.sub(r"[^\S\n]+", " ", text)
    text = re.sub(r"\n+", "\n", text).strip()
    text = re.sub(r"^(?:岗位|职位|招聘岗位|招聘职位|岗位名称|职位名称)\s*[：:]\s*", "", text)
    if not text:
        return []
    strong_parts = [
        _clean_job_name(part)
        for part in split_outside_brackets(text, {"\n", ";", "；"})
        if _clean_job_name(part)
    ]
    parts: list[str] = []
    for part in strong_parts or [text]:
        comma_parts = [
            _clean_job_name(item)
            for item in split_outside_brackets(part, {"、", ",", "，"})
            if _clean_job_name(item)
        ]
        complete_names = len(comma_parts) > 1 and all(
            _ends_like_role(item) for item in comma_parts
        )
        explicit_long_list = (
            len(comma_parts) >= 3
            and len(part) >= 20
            and all(2 <= len(item) <= 50 and _looks_like_job_name(item) for item in comma_parts)
        )
        if complete_names or explicit_long_list:
            parts.extend(item.removesuffix("等") for item in comma_parts)
        else:
            slash_parts = [
                _clean_job_name(item)
                for item in split_outside_brackets(part, {"/", "／"})
                if _clean_job_name(item)
            ]
            colon_position = min(
                (position for marker in (":", "：") if (position := part.find(marker)) >= 0),
                default=-1,
            )
            first_slash = min(
                (position for marker in ("/", "／") if (position := part.find(marker)) >= 0),
                default=-1,
            )
            explicit_slash_list = (
                len(slash_parts) >= 3
                and len(part) >= 8
                and not (0 <= colon_position < first_slash)
                and all(2 <= len(item) <= 50 and _looks_like_job_name(item) for item in slash_parts)
            )
            if explicit_slash_list:
                parts.extend(item.removesuffix("等") for item in slash_parts)
                continue

            whitespace_parts = [
                _clean_job_name(item)
                for item in split_outside_brackets(part, set(), whitespace=True)
                if _clean_job_name(item)
            ]
            all_have_cjk = all(re.search(r"[\u4e00-\u9fff]", item) for item in whitespace_parts)
            explicit_whitespace_list = (
                len(whitespace_parts) >= 2
                and all_have_cjk
                and (
                    len(whitespace_parts) >= 3
                    or all(_ends_like_role(item) for item in whitespace_parts)
                )
                and all(_looks_like_job_name(item) for item in whitespace_parts)
            )
            candidates = whitespace_parts if explicit_whitespace_list else [part]
            for candidate in candidates:
                parts.extend(split_repeated_strong_roles(candidate))
    result: list[str] = []
    fully_split_parts: list[str] = []
    for part in parts:
        fully_split_parts.extend(split_repeated_strong_roles(part))
    for part in fully_split_parts:
        if _looks_like_job_name(part) and part not in result:
            result.append(part)
    return result


def html_table_grid(table: Tag) -> list[list[str]]:
    """Expand HTML rowspan/colspan cells into a rectangular logical grid."""

    rows: list[list[str]] = []
    pending: dict[int, tuple[str, int]] = {}

    def consume_span(row: list[str], column: int) -> None:
        text, remaining = pending[column]
        while len(row) < column:
            row.append("")
        row.append(text)
        if remaining <= 1:
            del pending[column]
        else:
            pending[column] = (text, remaining - 1)

    for tr in table.find_all("tr"):
        row: list[str] = []
        column = 0
        cells = tr.find_all(["th", "td"], recursive=False)
        for cell in cells:
            while column in pending:
                consume_span(row, column)
                column += 1
            lines = [normalize_text(line) for line in cell.get_text("\n", strip=True).splitlines()]
            text = "\n".join(line for line in lines if line)
            try:
                rowspan = max(1, int(cell.get("rowspan", 1)))
                colspan = max(1, int(cell.get("colspan", 1)))
            except (TypeError, ValueError):
                rowspan, colspan = 1, 1
            for _ in range(colspan):
                while len(row) < column:
                    row.append("")
                row.append(text)
                if rowspan > 1:
                    pending[column] = (text, rowspan - 1)
                column += 1
        if pending:
            last_pending = max(pending)
            while column <= last_pending:
                if column in pending:
                    consume_span(row, column)
                else:
                    row.append("")
                column += 1
        if row:
            rows.append(row)
    return rows


def extract_table_positions(content_html: str) -> list[Position]:
    soup = BeautifulSoup(content_html or "", "html.parser")
    positions: list[Position] = []
    for table in soup.find_all("table"):
        rows = html_table_grid(table)
        header_row = -1
        job_index: int | None = None
        for index, row in enumerate(rows[:4]):
            located = _find_job_header(row)
            if located is not None:
                header_row, job_index = index, located
                break
        if job_index is None:
            continue
        headers = rows[header_row]
        major_index = _find_header(headers, MAJOR_HEADER_NAMES)
        education_index = _find_header(headers, EDUCATION_HEADER_NAMES)
        location_index = _find_header(headers, LOCATION_HEADER_NAMES)
        deadline_index = _find_header(headers, DEADLINE_HEADER_NAMES)

        def cell(row: Sequence[str], index: int | None) -> str:
            return normalize_text(row[index]) if index is not None and index < len(row) else ""

        for row in rows[header_row + 1 :]:
            if job_index >= len(row) or _find_job_header(row) is not None:
                continue
            names = split_job_names(row[job_index])
            for name in names:
                positions.append(
                    Position(
                        name=name,
                        education=cell(row, education_index),
                        major=cell(row, major_index),
                        location=cell(row, location_index),
                        deadline=cell(row, deadline_index),
                    )
                )
    return positions


def extract_labeled_positions(content_html: str) -> list[Position]:
    soup = BeautifulSoup(content_html or "", "html.parser")
    positions: list[Position] = []
    pattern = re.compile(
        r"^(?:\d+[.、)）]\s*)?(?:招聘)?(?:岗位|职位)(?:名称)?\s*[：:]\s*(.{2,160})$"
    )
    for tag in soup.find_all(["p", "li", "h2", "h3", "h4", "strong"]):
        text = normalize_text(tag.get_text("\n", strip=True))
        match = pattern.match(text)
        if not match:
            continue
        for name in split_job_names(match.group(1)):
            positions.append(Position(name=name))
    return positions


def deduplicate_positions(positions: Iterable[Position]) -> list[Position]:
    result: dict[str, Position] = {}
    for position in positions:
        key = normalize_text(position.name)
        if not key:
            continue
        existing = result.get(key)
        if existing is None:
            result[key] = Position(
                name=key,
                education=normalize_text(position.education),
                major=normalize_text(position.major),
                location=normalize_text(position.location),
                deadline=normalize_text(position.deadline),
            )
        else:
            for attr in ("education", "major", "location", "deadline"):
                if not getattr(existing, attr) and getattr(position, attr):
                    setattr(existing, attr, normalize_text(getattr(position, attr)))
    return list(result.values())


def extract_positions(detail: Mapping[str, Any]) -> list[Position]:
    content = normalize_text(detail.get("recruitmentContent"))
    table_positions = extract_table_positions(str(detail.get("recruitmentContent") or ""))
    if table_positions:
        return deduplicate_positions(table_positions)
    labeled_positions = extract_labeled_positions(str(detail.get("recruitmentContent") or ""))
    if labeled_positions:
        return deduplicate_positions(labeled_positions)
    api_title = normalize_text(detail.get("title"))
    if api_title:
        api_positions = [Position(name=name) for name in split_job_names(api_title)]
        if api_positions:
            return deduplicate_positions(api_positions)
    del content  # Explicitly leave the field blank when no reliable position was found.
    return [Position(name="未明确列出")]


def extract_submission(detail: Mapping[str, Any], content_html: str) -> tuple[str, str]:
    method = normalize_text(detail.get("resumeSubmissionMethod"))
    explicit_urls: list[str] = []
    for field_name in ("onlineUrl", "preachingUrls"):
        raw = detail.get(field_name)
        values: list[Any]
        if isinstance(raw, str) and raw.strip().startswith("["):
            try:
                decoded = json.loads(raw)
                values = decoded if isinstance(decoded, list) else [decoded]
            except json.JSONDecodeError:
                values = [raw]
        elif isinstance(raw, list):
            values = raw
        else:
            values = [raw]
        for value in values:
            url = normalize_url(value)
            if url and urlsplit(url).scheme in {"http", "https"} and url not in explicit_urls:
                explicit_urls.append(url)

    soup = BeautifulSoup(content_html or "", "html.parser")
    for anchor in soup.find_all("a", href=True):
        anchor_text = normalize_text(anchor.get_text(" ", strip=True))
        parent_text = normalize_text(anchor.parent.get_text(" ", strip=True)) if anchor.parent else anchor_text
        context = f"{anchor_text} {parent_text}"[:300]
        if not re.search(r"投递|网申|申请|报名|招聘官网|简历", context):
            continue
        url = normalize_url(html.unescape(anchor.get("href", "")))
        if url and urlsplit(url).scheme in {"http", "https"} and url not in explicit_urls:
            explicit_urls.append(url)

    plain_url_re = re.compile(r"https?://[A-Za-z0-9._~:/?#\[\]@!$&*+=%\-]+", re.IGNORECASE)
    for tag in soup.find_all(["p", "li", "div"], limit=500):
        text = normalize_text(tag.get_text(" ", strip=True))
        matches = plain_url_re.findall(text)
        if not matches:
            continue
        context_parts = [text]
        sibling = tag.previous_sibling
        while sibling is not None and len(context_parts) < 3:
            if isinstance(sibling, Tag):
                sibling_text = normalize_text(sibling.get_text(" ", strip=True))
                if sibling_text:
                    context_parts.append(sibling_text)
            sibling = sibling.previous_sibling
        context = " ".join(context_parts)[:500]
        if not re.search(r"投递|网申|申请|报名|招聘网|登录网址|简历", context):
            continue
        for raw_url in matches:
            url = normalize_url(html.unescape(raw_url.rstrip(".,;；，。")))
            if url and urlsplit(url).scheme in {"http", "https"} and url not in explicit_urls:
                explicit_urls.append(url)

    if not method:
        for tag in soup.find_all(["p", "li"], limit=300):
            text = normalize_text(tag.get_text(" ", strip=True))
            match = re.match(r"^(?:简历)?投递(?:方式|渠道)?\s*[：:]\s*(.{2,160})$", text)
            if match:
                method = normalize_text(match.group(1))
                break
    return method, explicit_urls[0] if explicit_urls else ""


def build_records(
    candidate: Candidate,
    detail: Mapping[str, Any],
    detail_url: str,
    link_status: str,
) -> list[dict[str, str]]:
    merged_detail = dict(candidate.raw)
    for key, value in detail.items():
        if value not in (None, "", []):
            merged_detail[key] = value
    content_html = str(merged_detail.get("recruitmentContent") or "")
    title = extract_recruitment_title(content_html, merged_detail.get("title"))
    positions = extract_positions(merged_detail)
    submission_method, submission_url = extract_submission(merged_detail, content_html)
    education = join_values(merged_detail.get("educationRequirementLabel"))
    major = join_values(merged_detail.get("targetMajor"))
    work_location = join_values(merged_detail.get("workLocation"))
    deadline = format_datetime(merged_detail.get("resumeEndTime"))
    recruitment_date = format_date(merged_detail.get("recruitmentDate"))
    publish_time = format_datetime(merged_detail.get("publishTime"))

    rows: list[dict[str, str]] = []
    for position in positions:
        is_onsite = candidate.source_type == "现场招聘"
        record = {
            "唯一ID": "",
            "首次获取时间": "",
            "最后更新时间": "",
            "信息类型": candidate.source_type,
            "公司名称": normalize_text(merged_detail.get("companyName")),
            "招聘标题": title,
            "岗位": normalize_text(position.name),
            "发布时间": publish_time,
            "招聘日期": recruitment_date,
            "工作地点": normalize_text(position.location) or work_location,
            "宣讲日期": recruitment_date if is_onsite else "",
            "宣讲开始时间": format_time(merged_detail.get("startTime")) if is_onsite else "",
            "宣讲结束时间": format_time(merged_detail.get("endTime")) if is_onsite else "",
            "宣讲地点": normalize_text(merged_detail.get("recruitmentLocation")) if is_onsite else "",
            "学历要求": normalize_text(position.education) or education,
            "专业要求": normalize_text(position.major) or major,
            "网申截止时间": format_datetime(position.deadline) or deadline,
            "投递方式": submission_method,
            "投递链接": submission_url,
            "详情链接": normalize_url(detail_url),
            "链接状态": link_status,
            "来源页面": normalize_url(candidate.source_page),
            "备注": "",
        }
        record["唯一ID"] = make_unique_id(record)
        rows.append(record)
    return rows


def verification_matches(candidate: Candidate, detail: Mapping[str, Any], page_text: str) -> bool:
    expected_company = normalize_text(candidate.raw.get("companyName"))
    actual_company = normalize_text(detail.get("companyName"))
    if expected_company and actual_company and expected_company == actual_company:
        return True
    expected_title = normalize_text(candidate.raw.get("title"))
    actual_title = normalize_text(detail.get("title"))
    if expected_title and actual_title and expected_title == actual_title:
        return True
    compact_page = normalize_text(page_text)
    return bool(expected_company and expected_company in compact_page)


def fetch_detail(client: UestcClient, candidate: Candidate) -> DetailResult:
    requested_url = candidate_detail_url(client.base_url, candidate)
    try:
        response = client.request("GET", requested_url)
    except FetchError as exc:
        records = build_records(candidate, {}, requested_url, "failed")
        return DetailResult(candidate, requested_url, "failed", records, str(exc))

    final_url = normalize_url(response.url)
    expected_host = urlsplit(client.base_url).netloc.lower()
    final_host = urlsplit(final_url).netloc.lower()
    if final_host != expected_host:
        records = build_records(candidate, {}, requested_url, "unverified")
        return DetailResult(candidate, requested_url, "unverified", records, "详情页跳转到非本站地址")

    detail = extract_nuxt_recruitment(response.text, candidate.announcement_id)
    if detail is None:
        try:
            payload = client.api_get(f"api/home/recruitment/{candidate.announcement_id}")
            api_detail = payload.get("data")
            detail = api_detail if isinstance(api_detail, dict) else {}
        except FetchError:
            detail = {}
    page_text = BeautifulSoup(response.text, "html.parser").get_text(" ", strip=True)
    status = "verified" if detail and verification_matches(candidate, detail, page_text) else "unverified"
    records = build_records(candidate, detail, final_url or requested_url, status)
    error = "" if status == "verified" else "详情页存在，但公司名称或标题未能充分匹配"
    return DetailResult(candidate, final_url or requested_url, status, records, error)


def fetch_details(
    client: UestcClient,
    candidates: Sequence[Candidate],
    workers: int,
) -> list[DetailResult]:
    if not candidates:
        return []
    ordered: list[DetailResult | None] = [None] * len(candidates)
    with ThreadPoolExecutor(max_workers=max(1, workers)) as executor:
        future_map = {
            executor.submit(fetch_detail, client, candidate): index
            for index, candidate in enumerate(candidates)
        }
        for future in as_completed(future_map):
            index = future_map[future]
            try:
                ordered[index] = future.result()
            except Exception as exc:  # One malformed detail must not stop other announcements.
                candidate = candidates[index]
                url = candidate_detail_url(client.base_url, candidate)
                ordered[index] = DetailResult(
                    candidate=candidate,
                    detail_url=url,
                    link_status="failed",
                    records=build_records(candidate, {}, url, "failed"),
                    error=f"{exc.__class__.__name__}: {exc}",
                )
    return [result for result in ordered if result is not None]


def candidate_detail_url(base_url: str, candidate: Candidate) -> str:
    """Return the canonical detail route dictated by the live site's type and ID."""

    encoded_id = encode_public_id(candidate.announcement_id)
    return normalize_url(urljoin(base_url, f"recruitment/{candidate.source_slug}/{encoded_id}"))


def _stringify_cell(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return normalize_text(value)


def read_csv_history(path: Path) -> list[dict[str, str]]:
    if not path.exists() or path.stat().st_size == 0:
        return []
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle)
                if not reader.fieldnames or "唯一ID" not in reader.fieldnames:
                    raise DataFileError(f"{path} 缺少唯一ID表头")
                rows = []
                for raw in reader:
                    row = {field: normalize_text(raw.get(field)) for field in CSV_FIELDS}
                    if row["唯一ID"]:
                        rows.append(row)
                return rows
        except UnicodeError as exc:
            last_error = exc
            continue
    raise DataFileError(f"无法安全读取 {path}: {last_error}")


def read_excel_history(path: Path) -> tuple[list[dict[str, str]], list[list[str]]]:
    if not path.exists() or path.stat().st_size == 0:
        return [], []
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise DataFileError(f"无法读取现有 Excel：{exc}") from exc
    try:
        if "招聘记录" not in workbook.sheetnames:
            raise DataFileError("现有 Excel 缺少“招聘记录”工作表")
        sheet = workbook["招聘记录"]
        header_values = [_stringify_cell(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_index = {name: index for index, name in enumerate(header_values) if name}
        if "唯一ID" not in header_index:
            raise DataFileError("现有 Excel 缺少唯一ID列")
        rows: list[dict[str, str]] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = {
                field: _stringify_cell(values[header_index[field]])
                if field in header_index and header_index[field] < len(values)
                else ""
                for field in CSV_FIELDS
            }
            if row["唯一ID"]:
                rows.append(row)

        updates: list[list[str]] = []
        if "更新记录" in workbook.sheetnames:
            update_sheet = workbook["更新记录"]
            update_headers = [
                _stringify_cell(cell.value)
                for cell in next(update_sheet.iter_rows(min_row=1, max_row=1))
            ]
            if update_headers[: len(UPDATE_FIELDS)] == UPDATE_FIELDS:
                for values in update_sheet.iter_rows(min_row=2, values_only=True):
                    if any(value not in (None, "") for value in values[: len(UPDATE_FIELDS)]):
                        updates.append([_stringify_cell(value) for value in values[: len(UPDATE_FIELDS)]])
        return rows, updates
    finally:
        workbook.close()


def load_history(csv_path: Path, excel_path: Path) -> tuple[list[dict[str, str]], list[list[str]], bool]:
    """Read both stores; CSV wins, while Excel-only IDs are recovered conservatively."""

    csv_rows = read_csv_history(csv_path)
    excel_rows: list[dict[str, str]] = []
    update_log: list[list[str]] = []
    excel_damaged = False
    try:
        excel_rows, update_log = read_excel_history(excel_path)
    except DataFileError:
        if not csv_rows:
            raise
        logging.warning("现有 Excel 无法读取，将从 CSV 安全重建。")
        excel_damaged = True
    combined: dict[str, dict[str, str]] = {row["唯一ID"]: row for row in csv_rows}
    recovered = False
    for row in excel_rows:
        if row["唯一ID"] not in combined:
            combined[row["唯一ID"]] = row
            recovered = True
    return list(combined.values()), update_log, recovered or excel_damaged


def _merge_one(existing: dict[str, str], incoming: Mapping[str, Any]) -> tuple[dict[str, str], list[str]]:
    merged = dict(existing)
    changes: list[str] = []
    for field_name in CSV_FIELDS:
        if field_name in IMMUTABLE_FIELDS:
            continue
        old = normalize_text(existing.get(field_name))
        new = normalize_text(incoming.get(field_name))
        if field_name == "链接状态":
            if LINK_STATUS_RANK.get(new, 0) <= LINK_STATUS_RANK.get(old, 0):
                continue
        elif not new:
            continue
        if new != old:
            merged[field_name] = new
            changes.append(f"{field_name}: {old or '空'} → {new}")
    return merged, changes


def merge_scraped_records(
    existing_rows: Sequence[Mapping[str, Any]],
    scraped_rows: Sequence[Mapping[str, Any]],
    timestamp: str,
) -> MergeResult:
    """Insert new IDs and update meaningful non-empty fields without duplicate rows."""

    rows = [{field: normalize_text(row.get(field)) for field in CSV_FIELDS} for row in existing_rows]
    by_id = {row["唯一ID"]: index for index, row in enumerate(rows) if row["唯一ID"]}
    current_seen: set[str] = set()
    new_records: list[dict[str, str]] = []
    updated_records: list[dict[str, str]] = []
    events: list[ChangeEvent] = []
    unchanged = 0

    for raw in scraped_rows:
        incoming = {field: normalize_text(raw.get(field)) for field in CSV_FIELDS}
        incoming["唯一ID"] = incoming["唯一ID"] or make_unique_id(incoming)
        unique_id = incoming["唯一ID"]
        if unique_id in current_seen:
            continue
        current_seen.add(unique_id)
        if unique_id not in by_id:
            incoming["首次获取时间"] = timestamp
            incoming["最后更新时间"] = timestamp
            by_id[unique_id] = len(rows)
            rows.append(incoming)
            new_records.append(incoming)
            events.append(
                ChangeEvent(timestamp, "新增", unique_id, incoming["公司名称"], incoming["岗位"], "新增招聘记录")
            )
            continue
        index = by_id[unique_id]
        merged, changes = _merge_one(rows[index], incoming)
        if changes:
            merged["首次获取时间"] = rows[index]["首次获取时间"]
            merged["最后更新时间"] = timestamp
            rows[index] = merged
            updated_records.append(merged)
            operation = "链接验证状态变化" if all(change.startswith("链接状态:") for change in changes) else "更新"
            events.append(
                ChangeEvent(
                    timestamp,
                    operation,
                    unique_id,
                    merged["公司名称"],
                    merged["岗位"],
                    "；".join(changes),
                )
            )
        else:
            unchanged += 1
    return MergeResult(rows, new_records, updated_records, events, unchanged)


def _temporary_path(target: Path, suffix: str) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    descriptor, name = tempfile.mkstemp(prefix=f".{target.name}.", suffix=suffix, dir=target.parent)
    os.close(descriptor)
    return Path(name)


def atomic_write_csv(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    temporary = _temporary_path(path, ".tmp")
    try:
        with temporary.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=CSV_FIELDS,
                extrasaction="ignore",
                lineterminator="\n",
            )
            writer.writeheader()
            for row in rows:
                writer.writerow({field: normalize_text(row.get(field)) for field in CSV_FIELDS})
            handle.flush()
            os.fsync(handle.fileno())
        validated = read_csv_history(temporary)
        if len(validated) != len(rows):
            raise DataFileError("CSV 临时文件记录数校验失败")
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _style_worksheet(sheet: Any, fields: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{max(1, len(rows) + 1)}"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    preferred = {
        "唯一ID": 18,
        "首次获取时间": 20,
        "最后更新时间": 20,
        "信息类型": 12,
        "公司名称": 28,
        "招聘标题": 40,
        "岗位": 26,
        "发布时间": 20,
        "招聘日期": 13,
        "工作地点": 28,
        "宣讲日期": 13,
        "宣讲开始时间": 14,
        "宣讲结束时间": 14,
        "宣讲地点": 26,
        "学历要求": 20,
        "专业要求": 32,
        "网申截止时间": 20,
        "投递方式": 30,
        "投递链接": 36,
        "详情链接": 42,
        "链接状态": 13,
        "来源页面": 36,
        "备注": 30,
    }
    for index, field_name in enumerate(fields, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = preferred.get(field_name, 18)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def atomic_write_excel(
    path: Path,
    rows: Sequence[Mapping[str, Any]],
    prior_updates: Sequence[Sequence[str]],
    events: Sequence[ChangeEvent],
) -> None:
    workbook = Workbook()
    records_sheet = workbook.active
    records_sheet.title = "招聘记录"
    records_sheet.append(EXCEL_FIELDS)
    for record in rows:
        records_sheet.append([normalize_text(record.get(field)) for field in EXCEL_FIELDS])
    _style_worksheet(records_sheet, EXCEL_FIELDS, rows)
    for row_index in range(2, records_sheet.max_row + 1):
        for field_name in URL_FIELDS:
            column = EXCEL_FIELDS.index(field_name) + 1
            cell = records_sheet.cell(row=row_index, column=column)
            value = normalize_url(cell.value)
            if value and urlsplit(value).scheme in {"http", "https"}:
                cell.hyperlink = value
                cell.style = "Hyperlink"

    updates_sheet = workbook.create_sheet("更新记录")
    updates_sheet.append(UPDATE_FIELDS)
    for values in prior_updates:
        updates_sheet.append(list(values)[: len(UPDATE_FIELDS)])
    for event in events:
        updates_sheet.append(
            [
                event.timestamp,
                event.operation,
                event.unique_id,
                event.company,
                event.job,
                event.description,
            ]
        )
    _style_worksheet(updates_sheet, UPDATE_FIELDS, [])
    for column, width in enumerate((20, 18, 18, 28, 26, 60), start=1):
        updates_sheet.column_dimensions[get_column_letter(column)].width = width

    temporary = _temporary_path(path, ".xlsx")
    try:
        workbook.save(temporary)
        workbook.close()
        validated_rows, _ = read_excel_history(temporary)
        if len(validated_rows) != len(rows):
            raise DataFileError("Excel 临时文件记录数校验失败")
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def markdown_safe(value: Any) -> str:
    return normalize_text(value).replace("|", "\\|")


def render_report_section(records: Sequence[Mapping[str, Any]], heading: str) -> list[str]:
    lines = [f"### {heading}", ""]
    if not records:
        lines.extend(["无。", ""])
        return lines
    for record in records:
        lines.extend(
            [
                f"- **公司名称：** {markdown_safe(record.get('公司名称'))}",
                f"  - 招聘标题：{markdown_safe(record.get('招聘标题')) or '未明确'}",
                f"  - 岗位：{markdown_safe(record.get('岗位')) or '未明确列出'}",
                f"  - 招聘/发布时间：{markdown_safe(record.get('招聘日期') or record.get('发布时间')) or '未提供'}",
                f"  - 工作地点：{markdown_safe(record.get('工作地点')) or '未提供'}",
            ]
        )
        if normalize_text(record.get("宣讲日期")):
            talk = " ".join(
                part
                for part in (
                    normalize_text(record.get("宣讲日期")),
                    normalize_text(record.get("宣讲开始时间")),
                    normalize_text(record.get("宣讲结束时间")),
                    normalize_text(record.get("宣讲地点")),
                )
                if part
            )
            lines.append(f"  - 宣讲时间和地点：{markdown_safe(talk)}")
        detail_url = normalize_url(record.get("详情链接"))
        lines.append(f"  - 详情链接：[{detail_url}]({detail_url})" if detail_url else "  - 详情链接：未获得")
        lines.extend([f"  - 链接验证状态：{markdown_safe(record.get('链接状态'))}", ""])
    return lines


def atomic_append_report(
    path: Path,
    timestamp: str,
    new_records: Sequence[Mapping[str, Any]],
    updated_records: Sequence[Mapping[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    existing = path.read_text(encoding="utf-8") if path.exists() else ""
    section: list[str] = []
    if not existing:
        section.extend(["# UESTC 招聘监控日报", ""])
    else:
        section.extend(["---", ""])
    section.extend(
        [
            f"## 运行时间：{timestamp}",
            "",
            f"- 新增招聘数量：{len(new_records)}",
            f"- 发生变化的招聘数量：{len(updated_records)}",
            "",
        ]
    )
    for source_type in ("现场招聘", "需求信息"):
        source_new = [record for record in new_records if record.get("信息类型") == source_type]
        source_updated = [record for record in updated_records if record.get("信息类型") == source_type]
        section.extend([f"## {source_type}", ""])
        section.extend(render_report_section(source_new, "新增"))
        section.extend(render_report_section(source_updated, "更新"))
    content = existing.rstrip() + ("\n\n" if existing else "") + "\n".join(section).rstrip() + "\n"
    temporary = _temporary_path(path, ".md")
    try:
        temporary.write_text(content, encoding="utf-8")
        if not temporary.read_text(encoding="utf-8").startswith("# UESTC 招聘监控日报"):
            raise DataFileError("Markdown 日报临时文件校验失败")
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def validate_data_consistency(csv_path: Path, excel_path: Path, expected_rows: int) -> None:
    csv_rows = read_csv_history(csv_path)
    excel_rows, _ = read_excel_history(excel_path)
    csv_ids = {row["唯一ID"] for row in csv_rows}
    excel_ids = {row["唯一ID"] for row in excel_rows}
    if len(csv_rows) != expected_rows or len(excel_rows) != expected_rows:
        raise DataFileError(
            f"记录数不一致：内存={expected_rows}, CSV={len(csv_rows)}, Excel={len(excel_rows)}"
        )
    if csv_ids != excel_ids or len(csv_ids) != expected_rows:
        raise DataFileError("CSV 与 Excel 的唯一ID集合不一致或存在重复")


def load_config(path: Path) -> dict[str, Any]:
    try:
        config = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise MonitorError(f"无法读取配置文件 {path}: {exc}") from exc
    required = {
        "timezone",
        "recent_days",
        "onsite_forward_days",
        "request_timeout",
        "request_interval_seconds",
        "max_retries",
        "list_page_size",
        "max_list_pages",
        "max_existing_rechecks_per_source",
        "detail_workers",
        "base_url",
    }
    missing = sorted(required - set(config))
    if missing:
        raise MonitorError(f"配置文件缺少字段：{', '.join(missing)}")
    try:
        ZoneInfo(str(config["timezone"]))
    except Exception as exc:
        raise MonitorError(f"无效时区：{config['timezone']}") from exc
    return config


def source_counts(records: Iterable[Mapping[str, Any]]) -> dict[str, int]:
    counts = {"现场招聘": 0, "需求信息": 0}
    for record in records:
        source = normalize_text(record.get("信息类型"))
        if source in counts:
            counts[source] += 1
    return counts


def print_summary(
    timestamp: str,
    sources: Mapping[str, SourceRun],
    total_rows: int,
    excel_rows: int,
    csv_written: bool,
    excel_written: bool,
    report_path: Path | None,
) -> None:
    print("UESTC Job Monitor")
    print("=================")
    print()
    print(f"运行时间：{timestamp}")
    print()
    for label in ("现场招聘", "需求信息"):
        run = sources[label]
        print(f"{label}：")
        if run.discovery_error:
            print(f"抓取失败：{run.discovery_error}")
        print(f"发现公告：{run.discovered}")
        print(f"本次检查公告：{run.selected}")
        print(f"拆分岗位记录：{run.parsed_rows}")
        print(f"新增：{run.new}")
        print(f"更新：{run.updated}")
        print()
    verified = sum(run.verified for run in sources.values())
    unverified = sum(run.unverified for run in sources.values())
    failed = sum(run.failed for run in sources.values())
    total_new = sum(run.new for run in sources.values())
    total_updated = sum(run.updated for run in sources.values())
    print("总计：")
    print(f"检查公告：{sum(run.selected for run in sources.values())}")
    print(f"拆分岗位记录：{sum(run.parsed_rows for run in sources.values())}")
    print(f"新增：{total_new}")
    print(f"更新：{total_updated}")
    print(f"详情链接 verified / unverified / failed：{verified} / {unverified} / {failed}")
    print(f"CSV记录总数：{total_rows}")
    print(f"Excel记录总数：{excel_rows}")
    print()
    print("Git：")
    print(f"数据发生变化：{'是' if total_new or total_updated else '否'}")
    if report_path:
        print(f"日报：{report_path.relative_to(ROOT)}")
    if total_new == 0 and total_updated == 0:
        print()
        print("今日未发现新增或变化的招聘信息。")
        print("历史数据未重复写入。")
        print("Excel未重新生成。" if not excel_written else "Excel因缺失或损坏已安全重建。")
    elif not csv_written or not excel_written:
        print("警告：数据有变化，但数据文件未全部写入。")


def run(config_path: Path) -> int:
    config = load_config(config_path)
    timezone = ZoneInfo(str(config["timezone"]))
    now = datetime.now(timezone)
    timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
    base_url = normalize_url(config["base_url"]) + "/"

    existing_rows, prior_updates, needs_repair = load_history(CSV_PATH, EXCEL_PATH)
    existing_urls = {
        normalize_url(row.get("详情链接"))
        for row in existing_rows
        if normalize_url(row.get("详情链接"))
    }
    client = UestcClient(config)
    sources = {
        "现场招聘": SourceRun("现场招聘"),
        "需求信息": SourceRun("需求信息"),
    }
    discovered_by_source: dict[str, list[Candidate]] = {"现场招聘": [], "需求信息": []}

    try:
        discovered_by_source["现场招聘"] = discover_onsite(
            client,
            base_url,
            now.date(),
            int(config["recent_days"]),
            int(config["onsite_forward_days"]),
        )
    except FetchError as exc:
        sources["现场招聘"].discovery_error = str(exc)
        logging.error("现场招聘抓取失败：%s", exc)
    try:
        cutoff = now.replace(tzinfo=None) - timedelta(days=int(config["recent_days"]))
        discovered_by_source["需求信息"] = discover_online(
            client,
            base_url,
            cutoff,
            int(config["list_page_size"]),
            int(config["max_list_pages"]),
        )
    except FetchError as exc:
        sources["需求信息"].discovery_error = str(exc)
        logging.error("需求信息抓取失败：%s", exc)

    if all(run.discovery_error for run in sources.values()):
        print("本次抓取失败，历史数据未修改。")
        return 2

    recheck_limit = int(config["max_existing_rechecks_per_source"])
    all_results: list[DetailResult] = []
    for source_type in ("现场招聘", "需求信息"):
        discovered = discovered_by_source[source_type]
        new_candidates: list[Candidate] = []
        existing_candidates: list[Candidate] = []
        for candidate in discovered:
            if candidate_detail_url(base_url, candidate) in existing_urls:
                existing_candidates.append(candidate)
            else:
                new_candidates.append(candidate)
        selected = new_candidates + existing_candidates[:recheck_limit]
        sources[source_type].discovered = len(discovered)
        sources[source_type].selected = len(selected)
        results = fetch_details(client, selected, int(config["detail_workers"]))
        all_results.extend(results)
        for result in results:
            setattr(sources[source_type], result.link_status, getattr(sources[source_type], result.link_status) + 1)

    scraped_rows: list[dict[str, str]] = []
    for result in all_results:
        # A transient detail failure must not create a new placeholder beside already
        # parsed rows for the same announcement, and must never downgrade old data.
        if result.link_status != "verified" and normalize_url(result.detail_url) in existing_urls:
            continue
        scraped_rows.extend(result.records)
    for source_type in ("现场招聘", "需求信息"):
        sources[source_type].parsed_rows = sum(
            1 for row in scraped_rows if row.get("信息类型") == source_type
        )

    merge = merge_scraped_records(existing_rows, scraped_rows, timestamp)
    new_counts = source_counts(merge.new_records)
    updated_counts = source_counts(merge.updated_records)
    for source_type in ("现场招聘", "需求信息"):
        sources[source_type].new = new_counts[source_type]
        sources[source_type].updated = updated_counts[source_type]

    data_changed = bool(merge.new_records or merge.updated_records)
    csv_written = False
    excel_written = False
    if data_changed or not CSV_PATH.exists() or needs_repair:
        atomic_write_csv(CSV_PATH, merge.rows)
        csv_written = True
    if data_changed or not EXCEL_PATH.exists() or needs_repair:
        atomic_write_excel(EXCEL_PATH, merge.rows, prior_updates, merge.events)
        excel_written = True
    report_path: Path | None = None
    if data_changed:
        report_path = REPORTS_DIR / f"{now.strftime('%Y-%m-%d')}.md"
        atomic_append_report(report_path, timestamp, merge.new_records, merge.updated_records)

    # If only one store was absent, ensure both are available before validating.
    if not CSV_PATH.exists():
        atomic_write_csv(CSV_PATH, merge.rows)
        csv_written = True
    if not EXCEL_PATH.exists():
        atomic_write_excel(EXCEL_PATH, merge.rows, prior_updates, merge.events)
        excel_written = True
    validate_data_consistency(CSV_PATH, EXCEL_PATH, len(merge.rows))
    excel_count = len(read_excel_history(EXCEL_PATH)[0])
    print_summary(
        timestamp,
        sources,
        len(merge.rows),
        excel_count,
        csv_written,
        excel_written,
        report_path,
    )
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--config",
        type=Path,
        default=DEFAULT_CONFIG_PATH,
        help="配置文件路径（默认 config/config.json）",
    )
    parser.add_argument("--verbose", action="store_true", help="输出调试日志")
    args = parser.parse_args()
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )
    try:
        return run(args.config.resolve())
    except (MonitorError, OSError) as exc:
        logging.error("%s", exc)
        print("本次抓取失败，历史数据未修改。")
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
